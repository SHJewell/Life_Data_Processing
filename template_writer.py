# -*- coding: utf-8 -*-
"""
Created on Fri Jan  1 12:23:24 2021

@author: Scott
template writer produces templates of log files to fill in

TODO
    Formatting still needs work
        colors and conditional formatting doesn't transfer
        Column widths need to be narrower

"""

import pandas as pd
import openpyxl as pxl
import os
    
activities = ["Friends",
             "Near Family",
             "Extended Family",
             "Steph",
             "Pets",
             "Sleep",
             "Work",
             "Prep",
             "Travel",
             "Errands",
             "Review",
             "Housekeeping",
             "Yardwork",
             "Reading",
             "Internet",
             "Games",
             "Leisure",
             "Distractions",
             "Brewing",
             "Sex",
             "Projects",
             "Eating",
             "Cooking",
             "Excercise",
             "Vacation",
             "Outdoors",
             "Transit"]



#def format_activies(xlsfile,acts):

    #for item in acts:
        
        
#weight tracker
def write_wt(year):
    
    col_titles = ["Weight", 
                  'Sleep', 
                  'Energy', 
                  'Mental', 
                  'Mood',
                  'Heart',
                  'Poop?',
                  'Drinks',
                  'coffee',
                  'Excercise?',
                  'Length',
                  'Intensity',
                  'Type']


    date0 = str(year) + '-01'
    dateN = str(year+1) + '-01'
    idx = pd.date_range(start=date0, end=dateN, freq='M')
    
    weight_sheets = {}
    
    for month in idx:
        
        num_days = pd.Period(month,freq='M').days_in_month
        strt = str(year) + "-" + str(month.month) + '-01'
        stp = str(year) + "-" + str(month.month) + "-" + str(num_days)
        
        days = pd.date_range(start=strt, end=stp, freq='D')
        days = days.date
        temp = pd.DataFrame(index=days, columns=col_titles)
        weight_sheets[month.month_name()] = temp
        
    # name = 'E:/Documents/Datasets/Life Data/' + str(year) + " sheets//" + str(year) + " weight tracker.xlsx"
    name = os.path.join(os.getcwd(), f'{path_parent}/{str(year)} sheets/{str(year)} weight tracker.xlsx')

    writer = pd.ExcelWriter(name, date_format='mm/dd')
    
    for mo_sheet in weight_sheets:

        weight_sheets[mo_sheet].to_excel(writer, sheet_name=mo_sheet, index_label="date", freeze_panes=(1, 1))
        
    writer.close()

#food tracker
def write_ft(year):
    
    col_titles = ['what', 'amount']

    date0 = str(year) + '-01'
    dateN = str(year+1) + '-01'
    idx = pd.date_range(start=date0, end=dateN, freq='M')
    
    monthly_sheets = {}
    
    for month in idx:
        
        num_days = pd.Period(month, freq='M').days_in_month
        strt = str(year) + "-" + str(month.month) + '-01'
        stp = str(year) + "-" + str(month.month) + "-" + str(num_days)
        
        row1 = []
        row2 = []
        weekN = 1
        months_sheet = {}

        for day in pd.date_range(start=strt, end=stp, freq='D'):
            
            if (day.dayofweek == 0) and (len(row1) != 0):
                
                temp = pd.DataFrame(columns=row1)
                temp.loc[0] = row2
                
                months_sheet["Week " + str(weekN)] = temp
                weekN += 1

                row1 = []
                row2 = []
            
            row1.append(day.date())
            row1.append(day.date())    #we need this twice
            row2.append(col_titles[0])
            row2.append(col_titles[1])
        
        temp = pd.DataFrame(columns=row1)
        temp.loc[0] = row2

        months_sheet["Week " + str(weekN)] = temp
        
        monthly_sheets[month.month_name()] = months_sheet
        
    for mo_sheet in monthly_sheets:


        #name = 'E:/Documents/Datasets/Life Data/' + str(year) + " sheets//" + str(year) + " " + mo_sheet + " food tracker.xlsx"
        name = os.path.join(os.getcwd(), f'{path_parent}/{str(year)} sheets/{str(year)} {mo_sheet} food tracker.xlsx')
        writer = pd.ExcelWriter(name, date_format='mm/dd')

        for week_sheet in monthly_sheets[mo_sheet]:
    
            monthly_sheets[mo_sheet][week_sheet].to_excel(writer, sheet_name=week_sheet, index=False)
        
        writer.close()

        '''
        this seems silly to save the workbook in pandas then reopen it in openpyxl,
        but at the time of writing it's simpler for formatting purposes, 
        especially for portion of the code that will probably only be used once a year
        '''
    
        wb = pxl.load_workbook(name)
                    
        for sheet in wb.sheetnames:
        
            if "Week" in sheet:
                
                cur_sheet = wb[sheet]
                n = 1
                
                while n < cur_sheet.max_column:
                        
                    cur_sheet.merge_cells(start_row=1, start_column=n, end_row=1, end_column=n+1)
                    n += 2
                                
        wb.save(name)

#daily log
def write_dl(year,activities):
    
    time = ['100']
    
    for n in range(1, 48):
        
        tminusone = str(time[n-1])
        
        if tminusone[-2:] == "00":
            
            time.append(str(int(tminusone) + 30))
            
        else:
            
            time.append(str(int(tminusone[:-2]) + 1) + "00")         
            
    date0 = str(year) + '-01'
    dateN = str(year+1) + '-01'
    idx = pd.date_range(start=date0, end=dateN, freq='M')
    
    daily_log = {}
    
    for month in idx:
        
        num_days = pd.Period(month,freq='M').days_in_month
        strt = str(year) + "-" + str(month.month) + '-01'
        stp = str(year) + "-" + str(month.month) + "-" + str(num_days)
        
        days = pd.date_range(start=strt, end=stp, freq='D')
        days = days.date
        temp = pd.DataFrame(index=days,columns=time)
        daily_log[month.month_name()] = temp
        
    # name = 'E:/Documents/Datasets/Life Data/' + str(year) + " sheets//" + "daily log " + str(year) + ".xlsx"
    name = os.path.join(os.getcwd(), f'{path_parent}/{str(year)} sheets/Daily log {str(year)}.xlsx')

    with pd.ExcelWriter(name, date_format='mm/dd', mode='w') as writer:
    
        for mnth in daily_log:

            daily_log[mnth].to_excel(writer, sheet_name=mnth, index_label="date", freeze_panes=(1, 1))

        #writer.to_excel()
    
    def add_column(sheet, column):
        
        sheet.insert_cols(1, 1)
        
        for rowy, value in enumerate(column, start=1):
            sheet.cell(row=rowy+1, column=1, value=value)
     
    #will need to     
    #template = "E:\Documents\Datasets\Life Data\Spreadsheets\daily template.xlsx"
    template = f'daily template.xlsx'
    
    x = pxl.load_workbook(template)
    y = pxl.load_workbook(name)
    
    wb = x['template'].conditional_formatting
    
    for month in y.sheetnames:
        
        y_cf = y[month].conditional_formatting
        
        for attr in list(dir(wb)):
            
            try:
                setattr(y_cf, attr, getattr(wb, attr))
            except:
                print(str(attr) + " cannot be written")
        
        if str(attr)[0] == "_":
            continue
        
        add_column(y[month], activities)

        #for col in list(x['template'])[0]:

        #print(col)

    y.save(name)

if __name__ == '__main__':

    year = 2024

    path_parent = './'

    if not os.path.isdir(os.path.join(os.getcwd(), path_parent, f'{year} sheets')):
        os.mkdir(os.path.join(os.getcwd(), path_parent, f'{year} sheets'))

    write_wt(year)
    write_ft(year)
    write_dl(year, activities)