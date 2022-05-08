# -*- coding: utf-8 -*-
"""
Created on Sat Aug  8 16:38:37 2020

@author: Scott

TODO:
    Current:
        Food master filled through 2020
    General
        Missing dates needs to be calculated
        Doesn't deal with duplicates in the food master file
        Writes skipped days as zeros, should be nans
    New Year:
        Maybe have the ability to add individual months instead of whole years
        Loop through months
            Loops will almost certainly be necessary, I don't see a way to do this without them
            Turn into ... dict? using date objects
            Convert to cals, etc per day

"""

import os
import openpyxl
import datetime
import csv
import string
import pickle
import pandas as pd
import re
#import tk
import utils

import logging

import pprint

logging.basicConfig(filename='E:\Documents\Datasets\Life Data\Logs\\food_concat_log.txt', level=logging.DEBUG)


class foodTracker:
    def __init__(self, path=None):
        self.today = datetime.datetime.today()
        self.file_path = path
        self.eating_log = {}        #???
        self.nutrition_calendar = pd.DataFrame
        self.food_master = pd.DataFrame
        self.reported_dates = set()
        self.missing_dates = set()
        self.year = None
        self.errors = []
        self.missing_food = []


    def import_month(self, filepath):
        '''
        Loops through month sheet fed by gen_new_log()
        :param filepath:
        :return None, updates self.nutrition_calendar:
        '''

        month = pd.read_excel(self.file_path + filepath, sheet_name=None)

        months_nutr = {}

        for name in month:

            if 'week' in name.lower():

                week = month[name]

                for dayN in range(0, len(week.columns), 2):

                    report_flag = False

                    days_nutrition = {'calories (kCal)':    0,
                                      'carbs (g)':          0,
                                      'fats (g)':           0,
                                      'fiber (g)':          0,
                                      'sugar (g)':          0
                                      }

                    try:
                        date = week.columns[dayN].date()
                    except AttributeError:
                        date = datetime.datetime.strptime(week.columns[dayN][:-2], '%Y-%m-%d %H:%M:%S').date()

                    day = week.iloc[1:, dayN:(dayN+2)].dropna()
                    #day.iloc[:, 0].apply(utils.word_processor).str.replace(' ', '')

                    for item in day.itertuples():

                        report_flag = True
                        food = utils.word_processor(item[1])

                        if food in self.food_master.index:

                            food_deets = self.food_master.loc[food, :].to_dict()

                            logging.debug('FOUND Date: ' + str(date) + ' ' + 'Item: ' + item[1] + ' as ' + food)

                            for item_key, amount in days_nutrition.items():

                                #problem with two instance of single item. Should be cleaned up on import
                                try:
                                    days_nutrition[item_key] = amount + item[2]*float(food_deets[item_key])

                                except TypeError:
                                    logging.debug(food + ' is a duplicate')
                                    self.errors.append(food + ' is a duplicate')
                                    days_nutrition[item_key] = amount + item[2]*float(food_deets[item_key][item[1]])

                        else:

                            self.missing_food.append([item[1], date])
                            logging.debug('NOT FOUND Date: ' + str(date) + ' ' + 'Item: |' + item[1] + '|')



                    if report_flag:
                        self.reported_dates.add(date)
                        months_nutr[date] = days_nutrition

        return months_nutr


    def export_logs(self, path):

        with open(path + '\\' + str(self.year) + '_food_port_log.txt', 'w+') as w:

            w.write('Errors\n')

            for line in self.errors:

                w.write(line)
                w.write('\n')

            w.write('Missing Dates\n')

            for date in self.missing_dates:

                w.write(str(date))
                w.write('\n')

        with open(path + '\\' + str(self.year) + '_missing_food.txt', 'w+') as w:

            for food, date in self.missing_food:

                w.write(str(date) + ': ' + food + '\n')


    def gen_new_log(self):
        '''
        Loops though months in year, and feeds existing food tracking sheets to import_month()
        :return:
        '''

        for file in os.listdir(self.file_path):

            if '~' in file or '#' in file:
                continue

            if 'food' in file.lower():

                file_year = utils.get_year(file)

                if self.year is None:

                    self.year = file_year

                elif self.year is not None and file_year != self.year:

                    self.errors.append("Years do not match!")
                    self.year = file_year

                month = self.import_month(file)

                try:
                    self.nutrition_calendar = pd.concat(
                        [self.nutrition_calendar, pd.DataFrame(month).transpose()])
                except TypeError:
                    self.nutrition_calendar = pd.DataFrame(month).transpose()


    def import_food_master(self, path):
        '''
        Import master food list
        updates self.food_master
        :param path:
        :return:
        '''

        self.food_master = pd.read_excel(path)
        self.food_master.index = self.food_master['food'].apply(utils.word_processor)

        if len(self.food_master.columns) > 8:
            self.food_master.drop(self.food_master.columns[8:], axis='columns', inplace=True)

        self.food_master.replace({'\n': '', '\r': ''}, regex=True, inplace=True)
        self.food_master.replace('', 0, inplace=True)
        self.food_master.fillna(0, inplace=True)
        self.food_master.drop_duplicates(subset='food', keep='first', inplace=True)


    def export_as_dat(self, path):

        self.nutrition_calendar.to_pickle(path + '\\' + str(self.year) + '_nutr_cal.dat')

    def export_as_csv(self, path):

        self.nutrition_calendar.to_csv(path + '\\' + str(self.year) + '_nutr_cal.csv')


if __name__ == '__main__':

    path = 'E:\Documents\Datasets\Life Data\\2022 sheets\\'
    master_path = 'E:\Documents\Datasets\Life Data\\Spreadsheets\\master food list.xlsx'

    foodlog = foodTracker(path=path)
    foodlog.import_food_master(master_path)
    foodlog.gen_new_log()
    foodlog.export_logs('E:\Documents\Datasets\Life Data\Logs')
    foodlog.export_as_csv('E:\Documents\Datasets\Life Data\Logs')
    #foodlog.export_as_dat('E:\Documents\Datasets\Life Data\Data files')

    # breakpoint()

# #==============================================================================
# #constants
#
# #path = "E:\\Documents\\Coding\\Life Data Processing\\2020 sheets"
# food_list_file = "master food list.xls"
#
# months = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
# titles = ['food','cals','carbs','fats','proteins','fiber','sugar','type']
#
# #JD adjuster for starting late in 2020
# JD_adjustement = 44
#
# master_calendar = {}
#
# blankdate = datetime.datetime(2020,1,1,0,0)
#
# #==============================================================================
# #Excel sheet opener
# #==============================================================================
#
# #returns a dictionary, where the worksheets are the keys
# #each sheet is a 2D nested list
#
# def open_xlsx(file):
#
#     wb = openpyxl.load_workbook(file)
#     sheets = wb.sheetnames
#
#     raw_sheets = {}
#
#     for sheet in sheets:
#
#         ws = wb[sheet]
#         cell_range = ws['A:N']
#         raw_data = []
#
#         for row in cell_range:
#             temp_row = []
#
#             for cell in row:
#
#                 temp_row.append(cell.value)
#
#             raw_data.append(temp_row)
#
#         raw_sheets[sheet] = raw_data
#
#     return raw_sheets
#
# #==============================================================================
# #generate a Julian date from a datetime.datetime object
#
# def gen_date(dt_obj):
#
#     leap_years = [2020, 2024, 2028, 2032, 2036, 2040, 2044, 2048, 2052, 2056, 2060, 2064, 2068, 2072, 2076, 2080, 2084, 2088, 2092, 2096]
#
#     normal_year_months = {
#         1 :   0,
#         2 :   31,
#         3 :   59,
#         4 :   90,
#         5 :   120,
#         6 :   151,
#         7 :   181,
#         8 :   212,
#         9 :   243,
#         10:   273,
#         11:   304,
#         12:   334
#         }
#
#     leap_year_months = {
#         1 :   0,
#         2 :   31,
#         3 :   60,
#         4 :   91,
#         5 :   121,
#         6 :   152,
#         7 :   182,
#         8 :   213,
#         9 :   244,
#         10:   274,
#         11:   305,
#         12:   335
#         }
#
#     yr = str(dt_obj.year)
#
#     # if yr in leap_years:
#     #     day = leap_year_months[dt_obj.month] + dt_obj.day
#     # else:
#     #     day = normal_year_months[dt_obj.month] + dt_obj.day
#
#     mon = str(dt_obj.month)
#     day = str(dt_obj.day)
#
#     if len(mon) == 1:
#         mon = "0" + mon
#
#     if len(day) == 1:
#         day = "0" + day
#
#     return yr + mon + day
#
#
# #==============================================================================
# #cleans up strings, removing leading and following spaces, carriage returns and
# #changing to lower case
#
# def word_processor(in_string):
#
#     mid = in_string.replace("\\n","")
#     mid = mid.replace("\n","")
#     mid = mid.replace("&","and")
#     mid = mid.replace(" ","")
#
#     mid = mid.translate(str.maketrans("","", string.punctuation))
#
#     if (mid in ["", None]) or (len(mid) == 0):
#         return
#
#     # while mid[0] == " ":
#
#     #     mid = mid[1:]
#
#     # while mid[-1] == " ":
#
#     #     mid = mid[0:-1]
#
#     if mid[-1] == "s":
#         mid = mid[:-1]
#
#     if mid in ["", None]:
#         return
#     else:
#         return mid.lower()
#
#
#
# #==============================================================================
# def day_processor(what,amount):
#
#     day = {}
#     #print(len(what),len(amount))
#
#     for n in range(2,len(what)):
#
#         #print(what[n],amount[n])
#
#
#         try:
#             item = word_processor(str(what[n]))
#         except:
#             print("STRING ERROR", "\'" + str(what[n]) + "\'")
#             continue
#
#         if item == "" or item == None:
#             continue
#
#         if amount[n] == None:
#             amount[n] = 0
#
#         if item in day:
#             day[item] += amount[n]
#
#         else:
#             day[item] = amount[n]
#
#     return day
#
# #==============================================================================
# #generate master sheet
#
# #generates adictionary for each JD, with the keys consistsing of the food name
# #and containing the amount
#
# def gen_master_calendar(current_master, new_month):
#     dummy_date = datetime.datetime(2020,1,1,0,0)
#
#     month = {}
#
#     for row in new_month:
#         if "Week" in row:
#
#             dayN = 0
#
#             #for dayN in range(len(new_month[row])):
#             while dayN < 14:
#                 day = new_month[row][dayN]
#
#
#                 if day == []:
#                     continue
#
#                 if type(day[0]) == type(dummy_date):
#                     #JD = gen_date(day[0])
#                     JD = day[0].date()
#                     #print(day[0],JD)
#
#                     how_much = new_month[row][dayN+1]
#                 #print(day[0],type(day[0]))
#                     what = new_month[row][dayN]
#
#                     day = day_processor(what,how_much)
#
#                     if len(day) > 1:
#                         month[JD] = day
#                     #print(day)
#
#                 dayN += 2
#
#     return month
#
# #Generate nutritional calendar=================================================
# def gen_nutr_cal(calendar,food):
#
#     missing = []
#     nutrition = {}
#     master = food.keys()
#     ed_master = []
#
#     for dish in master:
#
#         ed_master.append(word_processor(dish))
#
#     titles = ['cals','carbs','fats','proteins','fiber','sugar']
#
#     for day in calendar:
#
#         food_day = {
#               'cals'     :   0,
#               'carbs'    :   0,
#               'fats'     :   0,
#               'proteins' :   0,
#               'fiber'    :   0,
#               'sugar'    :   0
#               }
#
#         for c_item in calendar[day]:
#
#             itemp = word_processor(c_item)
#
#             if itemp == None or itemp == "none" or itemp == "":
#                 continue
#
#             if itemp not in ed_master:
#                 misdate = [c_item , day]
#                 missing.append(misdate)
#                 continue
#
#             for ingred in titles:
#
#                 food_ing = (food[itemp][ingred]).translate(dict.fromkeys(map(ord, string.whitespace)))
#
#                 if (food_ing == "") or (food_ing == None):
#                     food_ing = 0
#
#                 #print("Cal: \"", calendar[day][itemp], "\"",type(calendar[day][itemp]),"Food: \"",food_ing, "\"",type(food_ing))
#                 food_day[ingred] = food_day[ingred] + float(calendar[day][itemp])*float(food_ing)
#
#         nutrition[day] = food_day
#
#     return nutrition, missing
#
# #==============================================================================
#
# def gen_csv(data, name):
#
#     with open(name, 'w', newline='') as f:
#
#         writer = csv.writer(f, delimiter=",")
#
#         for row in data:
#
#             out = [row]
#
#             for item in data[row]:
#
#                 out.append(data[row][item])
#
#             writer.writerow(out)
#
# #==============================================================================
# #main
# #initialize variables
#
# ordered_months = {}
# raw_data = {}
#
# master_calendar = {}
#
# path = tkinter.filedialog.askopendirectory()
# files = os.listdir(path)
#
# #==============================================================================
# #food information
#
# master_file = path + "\\" + "master food list.csv"
#
# food_master = {}
#
# with open(master_file, newline='') as csvfile:
#     reader = csv.reader(csvfile)
#     for row in reader:
#
#         if row[0] in ['', 'food']:
#             continue
#
#         item = {}
#
#         for n in range(1,len(titles)):
#
#             #item[titles[n]] = word_processor(row[n])
#             item[titles[n]] = row[n]
#
#             food_master[word_processor(row[0])] = item
#             food_master["Name"] = item
#
#
# #pprint.pprint(food_master)
#
# for month in months:
#
#     for file in files:
#
#         if ".xlsx" not in file:
#             continue
#
#         if "#" in file:
#             continue
#
#         if month in file.lower():
#             ordered_months[n] = file
#             #files.remove(file)
#             n += 1
#
# for file in ordered_months:
#
#     filepath = path + "\\" + ordered_months[file]
#
#     if "xls" in filepath:
#         month = gen_master_calendar(master_calendar, open_xlsx(filepath))
#
#         master_calendar = {**master_calendar, **month}
#
#
# temp = gen_nutr_cal(master_calendar,food_master)
#
# nutr_cal = temp[0]
# missing_foods = temp[1]
# gen_csv(missing_foods, "missing food.csv")
# gen_csv(nutr_cal, "nutrition calendar.csv")
#
# #[date,calories,carbs,fats,proteins,fibs,sugar]
# pickle.dump(nutr_cal, open('nutritional_calendar.dat', "wb"))
#
#
# #pprint.pprint(check_missing_food(master_calendar, food_master))
#
# #pprint.pprint(master_calendar)