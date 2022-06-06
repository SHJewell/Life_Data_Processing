# -*- coding: utf-8 -*-
"""
Created on Sat Aug  8 16:38:57 2020

@author: Scott

TODO

"""

import openpyxl
import pandas as pd
import utils

class weightLog():
    def __init__(self):
        self.daily_acts = pd.DataFrame
        self.daily_ex = pd.DataFrame
        self.reported_dates = set()
        self.missing_dates = set()
        self.year = None
        self.errors = []
        self.path = None

    def read_log(self, path):
        '''
        TODO:
            Drop unnammed columns
            Rename columns for simplification sakes
        :param path:
        :return:
        '''

        self.path = path
        temp = pd.read_excel(path, sheet_name=None)
        emerge = []
        amerge = []

        for month in temp:

            month_df = temp[month].loc[:, ~temp[month].columns.str.contains('^Unnamed')]

            if len(month_df.columns) == 12:
                month_df.columns = ['date', 'weight', 'sleep', 'energy', 'mental', 'heart',
                                'poop', 'drinks', 'coffee', 'excercise', 'length', 'intensity']

                amerge.append(month_df)
                emerge.append(month_df.loc[:, ['date', 'excercise', 'length', 'intensity']].dropna(subset=['length']))

            elif len(month_df.columns) == 13:
                month_df.columns = ['date', 'weight', 'sleep', 'energy', 'mental', 'heart',
                                'poop', 'drinks', 'coffee', 'excercise', 'length', 'intensity', 'type']

                amerge.append(month_df)
                emerge.append(month_df.loc[:, ['date', 'excercise', 'length', 'intensity', 'type']].dropna(subset=['length']))

            else:
                self.errors.append(f'{month} has illegal number of columns!')

            # try:
            #     acts = month_df.loc[:, ['date', 'weight', 'sleep', 'energy', 'mental', 'heart', 'poop', 'drinks',
            #                             'coffee']]
            #     acts = acts[acts['date'].notna()]
            # except KeyError:
            #     self.errors.append(f'{month} has bad columns!')

        self.daily_acts = pd.concat(amerge)
        self.daily_acts.dropna(axis='index', subset=['date'], inplace=True)
        self.daily_acts.sort_values(by='date', ascending=False, inplace=True)

        self.daily_ex = pd.concat(emerge)
        self.daily_ex.fillna(method='backfill', inplace=True)
        self.daily_ex.sort_values(by='date', ascending=False, inplace=True)

        self.year = utils.get_year(path)

        return


    def ret_acts(self):

        return self.daily_acts


    def ret_ex(self):

        return self.daily_ex

    def ret_date(self):

        return self.year


#loc = "E:/Documents/Coding/Life Data Processing/2020 sheets"

if __name__ == '__main__':

    path = 'E:\Documents\Datasets\Life Data\\2020 sheets\Weight Tracker 2020.xlsx'
    #temp = pd.read_excel(path, sheet_name=None, engine='openpyxl')

    weight20 = weightLog()

    sheet = weight20.read_log(path)

    breakpoint()

#==============================================================================
#open xlsx, generic

# def open_xlsx(file):
    
#     wb = openpyxl.load_workbook(file)
#     sheets = wb.sheetnames
    
#     raw_sheets = {}
    
#     for sheet in sheets:
        
#         ws = wb[sheet]
#         cell_range = ws['A:N']
#         raw_data = []
        
#         for row in cell_range:   
#             temp_row = []
            
#             for cellN in range(len(row)):
                
#                 cell = row[cellN]
                
#                 temp_row.append(cell.value)
                
#             raw_data.append(temp_row)
        
#         raw_sheets[sheet] = raw_data
        
#     return raw_sheets

#==============================================================================
#open xlsx w/ pandas
    
# def data_from_xls(file_path):
#
#     raw = {}
#
#     sheets = pd.ExcelFile(file_path).sheet_names
#
#     for page in sheets:
#
#         temp = pd.read_excel(file_path,sheet_name=page)
#
#         #removes colmuns with empty names
#         raw[page] = temp.loc[:, ~temp.columns.str.contains("^Unnamed")]
#
#     return raw

#==============================================================================
#generate a datestring from a datetime.datetime object
    
# def date_string_gen(dt_obj):
#
#     out_str = (dt_obj.year)
#
#     month = (dt_obj.month)
#     day = (dt_obj.day)
#
#     new_date = datetime.date(out_str,month,day)
#
#     return new_date
#
# #==============================================================================
# #    MAIN
# #==============================================================================
#
# data = data_from_xls(filepath)
# header = data['January'].columns
# last_day = "20200000"
# master = {}
# days_acts = {}
#
# for month_name in data:
#
#     month = data[month_name]
#     dayN = 0
#
#     while dayN < month.shape[0]:
#
#         empty = True
#
#         if type(month.iloc[dayN,0]) != type(dt_object):
#
#             for ex in range(10,12):
#
#                 # if month.iloc[dayN, ex] == "no" or pd.isnull(month.iloc[dayN, ex]):
#                 #     break
#
#                 if not pd.isna(month.iloc[dayN,ex]):
#                     empty = False
#
#                 days_acts[header[ex]] = [month.iloc[dayN-1, ex], month.iloc[dayN, ex]]
#
#         else:
#             days_acts = {}
#
#             for itemN in range(1,len(header)-1):
#
#                 days_acts[header[itemN]] = month.iloc[dayN,itemN]
#
#             if not pd.isna(month.iloc[dayN,itemN]):
#                 empty = False
#
#             last_day = date_string_gen(month.iloc[dayN,0])
#
#         if not empty:
#             master[last_day] = days_acts
#
#         dayN += 1
#
# #pprint.pprint(master)
# #pprint.pprint(master[datetime.date(2020, 3, 23)])
#
# pickle.dump(master, open('weight_tracker.dat', "wb"))


